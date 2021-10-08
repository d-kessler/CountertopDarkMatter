import os
import csv
import openpyxl

from python.vars.paths_and_ids import name_id_manifest_path

# Ensuring that the current working directory is "CountertopDarkMatter"
while os.getcwd()[-20:] != "CountertopDarkMatter":
    os.chdir(os.path.join(".."))


def get_name_id_dict():
    """
    Load a dictionary of the form {'file_name': google_drive_id} from 'Name_ID.xlsx'.
    """
    name_id_manifest = ExcelUtils(name_id_manifest_path)
    wb, ws, first_empty_row = \
        name_id_manifest.wb, name_id_manifest.ws, name_id_manifest.first_empty_row
    name_id_dict = {}
    for row in range(2, first_empty_row):
        if ws.cell(row, 1).value is not None and ws.cell(row, 2).value is not None:
            name_id_dict[ws.cell(row, 1).value] = ws.cell(row, 2).value
    return name_id_dict


def fill_dict_from_dict(dict_to_fill, dict_containing_data, convert_types_to_string=None):
    """
    Pairs the keys of 'dict_to_fill' with the values paired with the same keys in 'dict_containing_data'.
    """
    if convert_types_to_string is None:
        convert_types_to_string = []
    for key in dict_containing_data.keys():
        if key in dict_to_fill.keys():
            if type(dict_containing_data[key]) in convert_types_to_string:
                dict_containing_data[key] = str(dict_containing_data[key])
            dict_to_fill[key] = dict_containing_data[key]
    return dict_to_fill


def verify_dict(dictionary, dict_name, correct_fieldnames, ignore_keys=None):
    """
    Checks whether 'dictionary' has all expected and no unexpected key-value pairs.
        dict_name: the name or type of 'dictionary' (str)
        correct_fieldnames: a list of fieldnames (str) that the keys of 'dictionary' should match
    """
    if ignore_keys is None:
        ignore_keys = []
    for key in dictionary.keys():
        if key in ignore_keys:
            continue
        if key not in correct_fieldnames:
            print(f'The {dict_name} dictionary has the added key {key}.')
            continue
        try:
            value = dictionary[key]
        except KeyError:
            print(f'The {dict_name} has no key {key}.')
            continue
        if value is None:
            print(f'In the {dict_name} dictionary, the value of the key {key} is None.')


def get_rows_dimension(rows_list):
    """
    Get the number of rows and columns in 'rows_list.' Used in 'write_rows' class methods.
    """
    if not rows_list:
        return rows_list, 0, 0
    if type(rows_list) == dict:
        return rows_list, 1, len(rows_list.keys())
    if (type(rows_list[0]) == list) \
            or (type(rows_list[0]) == dict) \
            or (type(rows_list[0]) == tuple) \
            or (type(rows_list[0]) == set):
        rows, cols = len(rows_list), len(rows_list[0])
        if rows == 1:
            rows_list = rows_list[0]
    else:
        rows, cols = 1, len(rows_list)
    return rows_list, rows, cols


class CsvUtils:
    def __init__(self, csv_path, fieldnames_list=None):
        """
        If the CSV does not exist, it is created with 'fieldnames_list' as its columns headers.
        If the CSV does exist and fieldnames_list is not given, it is read from the existing CSV.
            csv_path: file path to the CSV file, existing or to-be-created
            fieldnames_list: list of column headers
        """
        self.csv_path = csv_path
        if fieldnames_list:
            self.fieldnames_list = fieldnames_list
        elif os.path.exists(csv_path):
            self.fieldnames_list = self.read_rows(0, 0)[0]
        else:
            self.fieldnames_list = []
        if not os.path.exists(self.csv_path):
            self.create_csv()
        try:
            existing_fieldnames = self.read_rows(0, 0)[0]
        except IndexError:
            existing_fieldnames = []
        try:
            assert self.fieldnames_list == existing_fieldnames
        except AssertionError:
            need_new = input(f'Existing and passed fieldnames for {csv_path} do not match.'
                             f'\n\t The existing fieldnames were {existing_fieldnames}.'
                             f'\n\t The fieldnames passed were: {self.fieldnames_list}.'
                             f'\n\t Would you like to enter new fieldnames? [y/n]: ')
            if need_new == 'y':
                self.fieldnames_list = eval(input('Enter a new list of fieldnames: '))

    def create_csv(self):
        """
        Create a CSV at self.csv_path with the given columns headers.
        """
        with open(self.csv_path, 'w', newline='') as f:
            csv_writer = csv.DictWriter(f, fieldnames=self.fieldnames_list)
            csv_writer.writeheader()

    def write_rows(self, rows_list, dict_writer=False):
        """
        Write rows into self.csv_path. If 'dict_writer' is False, 'rows_list' must be a single- or
        multi-dimensional list of lists. If dict_writer is True, rows_list may be a dictionary or
        list of dictionaries with keys the columns headers of self.csv_path.
            rows_list: Single- or multi-dimensional list of csv rows to write.
        """
        rows_list, rows_list_rows, rows_list_cols = get_rows_dimension(rows_list)
        with open(self.csv_path, 'a', newline='') as f:
            if dict_writer is False:
                csv_writer = csv.writer(f)
            else:
                try:
                    assert (len(self.fieldnames_list) == rows_list_cols)
                except AssertionError:
                    print(f"CsvUtils {self.csv_path}, write_rows: fieldname # columns and row # columns do not match.")
                csv_writer = csv.DictWriter(f, fieldnames=self.fieldnames_list)
            if rows_list_rows == 1:
                csv_writer.writerow(rows_list)
            elif rows_list_rows > 1:
                csv_writer.writerows(rows_list)

    def read_rows(self, start_row=None, end_row=None, dict_reader=False):
        """
        Read rows of self.csv_path. If 'start_row' is not given, the reading begins at the first row;
        if 'end_row' is not given, the reading continues until the last row. If 'dict_reader' is false,
        the rows are returned as a list of lists; if dict_reader is true, a list of dicts is returned
        with the columns headers as keys.
            start_row: number of the first CSV row to be read
            end_row: number of the last CSV row to be read
            dict_reader: 'False' to return a list of lists, 'True' to return a list of dicts
        """
        if start_row is None:
            start_row = 0
        rows = []
        with open(self.csv_path, 'r') as f:
            if dict_reader is False:
                reader = csv.reader(f)
            else:
                reader = csv.DictReader(f)
            for i, row in enumerate(reader):
                if end_row:
                    if start_row <= i <= end_row:
                        rows.append(row)
                else:
                    if start_row <= i and row:
                        rows.append(row)
                    elif not row:
                        break
        return rows

    def find_row(self, identifier, column_number=None, column_header=None):
        with open(self.csv_path, 'r') as f:
            if column_number:
                csv_reader = csv.reader(f)
                for row in csv_reader:
                    if type(identifier)(row[column_number]) == identifier:
                        return row
            elif column_header:
                csv_reader = csv.DictReader(f)
                for row in csv_reader:
                    if type(identifier)(row[column_header]) == identifier:
                        return row
            else:
                print('Either column_number or column_header must be passed.')
                return None

    def clear(self):
        """
        Clear all content in self.csv_path, while keeping its columns headers.
        """
        os.remove(self.csv_path)
        self.create_csv()


class ExcelUtils:
    def __init__(self, excel_file_path, fieldnames_list=None):
        """
        If the excel file does not exist, it is created with 'fieldnames_list' as its columns headers.
        If the excel file does exist and fieldnames_list is not given, it is read from the existing excel file;
        if fieldnames_list was given and the existing excel file has no column headers, they are written in.
        The first empty row in the excel file is found; its row number is stored in self.first_empty_row.
            csv_path: file path to the CSV file, existing or to-be-created
            fieldnames_list: list of column headers
        """
        self.excel_file_path = excel_file_path
        self.fieldnames_list = fieldnames_list
        if not os.path.exists(excel_file_path):
            self.create_excel()
        self.wb, self.ws = self.configure_excel()
        self.first_empty_row = self.get_first_empty_row()
        if self.first_empty_row == 1:
            self.write_fieldnames()
        if fieldnames_list:
            self.fieldnames_list = fieldnames_list
            try:
                assert self.fieldnames_list == self.read_rows(1, 1)
            except AssertionError:
                need_new = input(f'Existing and passed fieldnames for {excel_file_path} do not match.'
                                 f'\n\t The existing fieldnames were {self.read_rows(1, 1)}.'
                                 f'\n\t The fieldnames passed were: {self.fieldnames_list}.'
                                 f'\n\t Would you like to enter new fieldnames? [y/n]: ')
                if need_new == 'y':
                    self.fieldnames_list = eval(input('Enter a new list of fieldnames: '))
        else:
            self.fieldnames_list = self.read_rows(1, 1)
        self.fieldname_columns_dict = self.get_fieldname_columns_dict()

    def create_excel(self):
        """
        Create an excel file at self.excel_file_path with the given column headers.
        """
        wb = openpyxl.Workbook()
        if self.fieldnames_list:
            ws = wb.active
            for col, val in enumerate(self.fieldnames_list, start=1):
                ws.cell(row=1, column=col).value = val
        wb.save(self.excel_file_path)

    def write_fieldnames(self):
        """
        Write the given column headers into self.excel_file_path.
        """
        for col, val in enumerate(self.fieldnames_list, start=1):
            self.ws.cell(row=1, column=col).value = val
        self.wb.save(self.excel_file_path)

    def configure_excel(self, ws=None):
        """
        Load 'wb' (workbook) and 'ws' (worksheet) openpyxl instances
        for the excel file that was given or created in __init__.
        If the name of a worksheet is not passed, the active (first)
        worksheet is used.
        """
        wb = openpyxl.load_workbook(filename=self.excel_file_path)
        if ws is None:
            ws = wb.active
        return wb, ws

    def get_first_empty_row(self):
        """
        Getting the first empty row of self.ws.
        """
        # Accounting for known bug in ws.max_row
        if not self.ws.cell(self.ws.max_row, 1).value:
            for row in range(self.ws.max_row, 0, -1):
                if self.ws.cell(row, 1).value:
                    return row + 1
            return 1
        return self.ws.max_row + 1

    def get_fieldname_columns_dict(self):
        fieldname_columns_dict = {}
        for i, fieldname in enumerate(self.fieldnames_list):
            fieldname_columns_dict[fieldname] = i
        return fieldname_columns_dict

    def read_rows(self, start_row=None, end_row=None):
        """
        Read rows of self.ws. If 'start_row' is not given, the reading begins at the first row;
        if 'end_row' is not given, the reading continues until the last row.
            start_row: number of the first excel worksheet row to be read
            end_row: number of the last excel worksheet row to be read
        """
        if start_row is None:
            start_row = 0
        if end_row is None:
            end_row = self.get_first_empty_row()
        rows = []
        if type(self.ws[start_row:end_row][0]) != tuple:
            return [c.value for c in self.ws[start_row:end_row]]
        for row in self.ws[start_row:end_row]:
            row_values = [c.value for c in row]
            if row_values != [None] * len(row_values):
                rows.append(row_values)
        return rows

    def write_rows(self, rows_list, starting_row=None, dict_writer=False):
        """
        Write rows into self.ws, starting on row 'starting_row'.
        'rows_list' is a single- or multi-dimensional list.
            rows_list: list (or list of lists) to write
            start_row: row number on which to start writing
        """
        rows_list, rows_list_rows, rows_list_cols = get_rows_dimension(rows_list)
        try:
            assert (len(self.fieldnames_list) == rows_list_cols)
        except AssertionError:
            print(f"ExcelUtils {self.excel_file_path}, write_rows: fieldname # columns and row # columns do not match.")
        if starting_row is None:
            sRow = self.get_first_empty_row()
        else:
            sRow = starting_row
        for row in range(rows_list_rows):
            if dict_writer is True:
                for fieldname in self.fieldname_columns_dict.keys():
                    col = self.fieldname_columns_dict[fieldname]
                    try:
                        cell_value = rows_list[row][fieldname]
                    except KeyError:
                        # Entered if `rows_list' only a dict (not a list of dicts)
                        cell_value = rows_list[fieldname]
                    try:
                        self.ws.cell(row=(sRow + row), column=(col + 1)).value = cell_value
                    except ValueError:
                        self.ws.cell(row=(sRow + row), column=(col + 1)).value = str(cell_value)
            else:
                for col in range(rows_list_cols):
                    if rows_list_rows == 1:
                        cell_value = rows_list[col]
                    else:
                        cell_value = rows_list[row][col]
                    try:
                        self.ws.cell(row=(sRow + row), column=(col + 1)).value = cell_value
                    except ValueError:
                        self.ws.cell(row=(sRow + row), column=(col + 1)).value = str(cell_value)
        self.wb.save(self.excel_file_path)

    def clear(self):
        self.wb.remove(self.ws)
        self.create_excel()
        self.wb, self.ws = self.configure_excel()
