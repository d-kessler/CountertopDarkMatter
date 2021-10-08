import os
import cv2
import numpy as np
from openpyxl.utils import get_column_letter

from python.utils.file_utils import make_folder
from python.utils.git_utils import push_files_to_GitHub
from python.utils.misc_utils import get_numerical_class_vars
from python.google_drive_folder.google_drive import GoogleDriveUtils
from python.utils.zooniverse_utils import ZooniverseUtils, upload_subjects_to_zooniverse
from python.utils.ellipse_utils import ellipse_eq_lhs, ellipse_extrema, draw_dashed_ellipse
from python.utils.csv_excel_utils import CsvUtils, ExcelUtils, fill_dict_from_dict, verify_dict

from python.vars.fieldnames import marking_fieldnames
from python.vars.project_info import marking_subject_set_id
from python.vars.subject_parameters import marking_edge_buffer_pixels, marking_image_enlarge_factor, \
    marking_border_gaps_angular_extent, marking_border_segments_angular_extent, marking_border_opacity, \
    marking_border_color, marking_border_thickness
from python.vars.paths_and_ids import marking_subjects_folder, fetched_images_folder, marking_manifest_path, \
    marking_manifest_csv_path, marking_csv_path, cleaned_classifications_manifest_csv_path, manifests_folder_drive_id, \
    manifests_csv_folder_drive_id, marking_folder_drive_id

# Ensuring that the current working directory is "CountertopDarkMatter"
while os.getcwd()[-20:] != "CountertopDarkMatter":
    os.chdir("..")

"""
(kSWAP instance).users[(user ID)] is an instance of the 'User' class, having attributes:
    user_id, classes, k = len(classes), gamma, user_default, user_score, confusion_matrix, and
    history = [(subject ID), ('user_score')]
        Example: [201, {"0": [0.6, 0.4], "1": [0.3, 0.7]}]
                    (  {"0": [True Negative, False Positive],
                        "1": [False Negative, True Positive]  )
(kSWAP instance).subjects[(subject ID)] is an instance of the 'Subject' class, having attributes:
    subject_id, score, classes, gold_label, epsilon, retired (boolean), retired_as, seen, and
    history = [(classification ID), (user ID), ('user_score'), (submitted classification), (subject score)]
        Example: [1001, 101, {"0": [0.6, 0.4], "1": [0.3, 0.7]}, 1, {"0": 0.35, "1": 0.65}]
"""


class PromoteSubjects:
    marking_metadata_row_template = dict((fieldname, None) for fieldname in marking_fieldnames)

    def __init__(self, swap, positive_prior, promotion_threshold, clear_folders=False):
        """
        Initializing , `subjects' (kSWAP subject instances), `promotion_threshold' (positive probability
        required for a feature to be added to the second workflow), Csv/ExcelUtil objects for
        `cleaned_classifications.csv', `marking_subjects.csv', and `Marking_Manifest.xlsx'.
            swap: kSWAP instance. Only used to retrieve kSWAP subject instances for the
                  Zooniverse subjects whose classifications have been analyzed
            positive_prior: the prior probability for an image to contain a meltpatch (ie. to be `positive')
            promotion_threshold: the ratio of posterior and prior positive probabilities required for features' promotion
        """
        self.subjects = list(swap.subjects.values())
        self.positive_prior = positive_prior
        self.promotion_threshold = promotion_threshold * positive_prior
        if clear_folders is True:
            self.clear_folders()
        self.cleaned_classifications_csv = CsvUtils(cleaned_classifications_manifest_csv_path)
        self.marking_csv = CsvUtils(marking_csv_path, fieldnames_list=marking_fieldnames)
        self.marking_manifest = ExcelUtils(marking_manifest_path, fieldnames_list=marking_fieldnames)
        self.marking_manifest_csv = CsvUtils(marking_manifest_csv_path, self.marking_manifest.fieldnames_list)
        self.gd = GoogleDriveUtils()

    def run(self):
        """
        Retrieving from `Marking_Manifest' a list of all classification IDs (integers) that 
        previously contributed to the promotion of a marking.
        """
        already_promoted_classification_ids = self.get_already_promoted_classification_ids()
        """
        Getting a list of `eligible' subjects, whose positive probability surpasses the threshold
        """
        eligible_subjects = self.get_subjects_past_positive_threshold()
        """
        Getting a dictionary with key-value pairs: 
            `eligible_subject (kSWAP instance)': [classification IDs where a marking was made]
        """
        eligible_subjects_marking_classification_ids = self.get_subjects_marking_classification_ids(eligible_subjects)
        """
        Deleting from classification ID lists in `eligible_subjects_marking_classification_ids' 
        the IDs that correspond to a marking that has already been promoted.
        """
        eligible_subjects_marking_classification_ids = self.remove_subjects_already_promoted_marking_ids(
            eligible_subjects_marking_classification_ids, already_promoted_classification_ids)
        """
        Stopping here if all marking IDs correspond to markings already promoted.
        """
        if list(eligible_subjects_marking_classification_ids.values()) == \
                [[]] * len(eligible_subjects_marking_classification_ids.values()):
            print('No markings to promote.')
            return
        """
        Replacing classification ID lists in `eligible_subjects_marking_classification_ids' with 
        lists whose elements are dictionaries containing markings' classification ID and parameters.
        The resulting dictionary has key value pairs:
            `eligible_subject (kSWAP instance)': [
                dictionaries like {"classification_id": , "cx": , "cy": , "xdim": , "ydim": , "angle": }]
        """
        eligible_subjects_markings_parameters = self.get_subjects_markings_parameters(
            eligible_subjects_marking_classification_ids)
        """
        Transforming the list of marking_dicts associated with each subject in `subjects_markings_parameters'
        into a list of tuples of markings_dicts that are related in the following manner: each marking_dict
        in a tuple corresponds to an ellipse that either contains the center of another marking_dict's ellipse
        or falls within the area occupied by another marking_dict's ellipse.
        """
        subjects_similar_markings = self.get_subjects_similar_markings(eligible_subjects_markings_parameters)
        """
        Getting a dictionary with key-value pairs
            'subject_id': [dictionaries corresponding to granite features whose individually-calculated 
                           probability of being `positive' surpasses the promotion threshold, formatted like
                            {'positive_probability': , 'subject_history': , 'similar_marking_dicts': } ]
        """
        subjects_features_to_promote = self.get_subjects_features_to_promote(subjects_similar_markings)
        """
        Creating and storing in `feature_to_promote_metadata_dicts' dictionaries of to-be-promoted features' metadata
        in the proper form for writing rows in the marking CSV and manifest
        """
        feature_to_promote_metadata_dicts = self.get_feature_to_promote_metadata_dicts(subjects_features_to_promote)
        """
        Retiring from Zooniverse experiment subjects in which there are promoted features.
        """
        self.retire_promoted(feature_to_promote_metadata_dicts)
        """
        Getting the names of all images in which the to-be-promoted features are contained.
        """
        needed_image_names = self.get_needed_image_names(feature_to_promote_metadata_dicts)
        """
        Adding all needed images to `fetched_images_folder'. Fetching local images where possible;
        downloading images from Google Drive if need be.
        """
        self.gd.fetch_needed_images(needed_image_names, search_second_folders=True)
        """
        Cropping the to-be-promoted features in the original (non-resized, higher-resolution) images 
        and drawing the `consensus marking', whose parameters are the average parameters of the
        markings made on the feature, using the to-be-promoted features' metadata 
        and class variables pertaining the the consensus markings' continuity, opacity, 
        color, and thickness as well as the final marking images' number of buffer pixels
        between the consensus marking and image edge and factor by which the to-be-promoted
        features' cropped images are to be enlarged.
        """
        self.create_marking_images(feature_to_promote_metadata_dicts)
        """
        Writing to-be-promoted features' metadata into the marking CSV and manifest.
        """
        self.write_metadata(feature_to_promote_metadata_dicts)
        """
        Uploading marking subjects to Zooniverse and Google Drive; uploads manifests to Github and Google Drive.
        """
        self.upload()

    def get_already_promoted_classification_ids(self):
        """
        Retrieves from `Marking_Manifest' a list of classification IDs (integers) that contributed
        to the promotion of a marking.
        """
        ws = self.marking_manifest.ws
        col_letter = 'V'
        for col in ws.iter_cols(1, ws.max_column):
            if col[0].value == "#classification_ids":
                col_letter = get_column_letter(col[0].column)
                break
        already_promoted = []
        for cell in ws[col_letter][1:]:
            cl_ids = eval(cell.value)
            for cl_id in cl_ids:
                if cl_id not in already_promoted:
                    already_promoted.append(int(cl_id))
        return already_promoted

    def get_subjects_past_positive_threshold(self):
        """
        Returns a list of subjects whose positive probabilities surpass the promotion threshold.
        """
        subjects_past_positive_threshold = []
        for subject in self.subjects:
            # Get subject's gold labels (-1: non-training, 0: negative, 1: positive)
            gold_label = subject.gold_label
            # Ignore training subjects
            if gold_label in [0, 1]:
                continue
            # Get subject's probability of being 'positive' (ie. of containing a melt-patch)
            positive_probability = subject.score['1']
            if positive_probability > self.promotion_threshold:
                subjects_past_positive_threshold.append(subject)
        return subjects_past_positive_threshold

    @staticmethod
    def get_subjects_marking_classification_ids(subjects):
        """Returns a dictionary of the form: 
            `subject instance': [classification IDs where a marking was made]
        ...
            subjects: list of kSWAP subject instances
        """
        subjects_marking_classification_ids = dict((subject, []) for subject in subjects)
        for subject in subjects:
            subject_history = subject.history
            classifications = np.array([sh[-2] for sh in subject_history][1:])  # ignore placeholder first entry
            marking_indices = list(np.where(classifications == 1)[0] + 1)  # `+ 1' to account for ignoring placeholder
            for marking_index in marking_indices:
                classification_id = subject_history[marking_index][0]
                subjects_marking_classification_ids[subject].append(classification_id)
        return subjects_marking_classification_ids

    @staticmethod
    def remove_subjects_already_promoted_marking_ids(eligible_subjects_marking_classification_ids,
                                                     already_promoted_classification_ids):
        """
        Deletes from classification ID lists in `eligible_subjects_marking_classification_ids' 
        the IDs that correspond to a marking that has already been promoted.
            eligible_subjects_marking_classification_ids: a dictionary with key-values:
                `eligible_subject (kSWAP instance)': [classification IDs where a marking was made]
            already_promoted_classification_ids: a list of classification IDs corresponding to markings
                                                 that have already been promoted
        """
        for subject in eligible_subjects_marking_classification_ids.keys():
            eligible_subjects_marking_classification_ids[subject] = \
                [ID for ID in eligible_subjects_marking_classification_ids[subject]
                 if ID not in already_promoted_classification_ids]
        return eligible_subjects_marking_classification_ids

    def get_subjects_markings_parameters(self, subjects_marking_classification_ids):
        # Iterate through subjects
        for subject in subjects_marking_classification_ids.keys():
            classification_ids = subjects_marking_classification_ids[subject]
            # Iterate through the markings that were made on the subject
            marking_dicts = []
            for ID in classification_ids:
                # Get the row of the cleaned classification CSV corresponding with this classification
                classification_row = self.cleaned_classifications_csv.find_row(ID, column_header='classification_id')
                # Append dictionary containing the marking's classification ID and parameters to `marking_dicts'
                marking_dicts.append({"classification_id": ID,
                                      "cx": float(classification_row['x']),
                                      "cy": float(classification_row['y']),
                                      "xdim": float(classification_row['rx']),
                                      "ydim": float(classification_row['ry']),
                                      "angle": -float(classification_row['angle'])})
                # Note: Zooniverse angles go CCW, while cv2 angles go CW
            # In `subjects_marking_classification_ids', replace the list of marking classification IDs with a
            # list of `marking_dicts', containing the classification ID and parameters
            subjects_marking_classification_ids[subject] = marking_dicts
        return subjects_marking_classification_ids

    @staticmethod
    def get_subjects_similar_markings(subjects_markings_parameters):
        """
        Transforms the list of marking_dicts associated with each subject in `subjects_markings_parameters'
        into a list of tuples of markings_dicts that are related in the following manner: each marking_dict
        in a tuple corresponds to an ellipse that either contains the center of another marking_dict's ellipse
        or falls within the area occupied by another marking_dict's ellipse.
        Note: any particular marking_dict in a tuple need only satisfy one of these two conditions with
        respect to ONE other marking_dict in the tuple; tuples may contain marking_dicts not related in
        these ways, so long as they are connected by intermediary markings so that they both belong to
        the same string of related markings.
            subjects_markings_parameters: a dictionary like
                `subject_id': [dictionaries like {"classification_id": , "cx": , "cy": , "xdim": , "ydim": , "angle": }]
        """

        def hashDict(dictionary):
            """
            Returns a "hashable" form of the given dictionary, such that it may be stored in a set.
            """
            return tuple(dictionary.items())

        def unhashDict(hashedDict):
            """
            Reverts a dictionary hashed by "hashDict" to its original form.
            """
            dictionary = {}
            for item in hashedDict:
                dictionary[item[0]] = item[1]
            return dictionary

        def reduce_list_of_sets(list_of_sets):
            """
            Combines all sets in `list_of_sets' with overlapping elements into a single set,
            returning a list of disjoint sets.
            """
            reduced_list_of_sets = []
            for s in list_of_sets:
                s_unions = s
                outside_s = []
                for r in reduced_list_of_sets:
                    if s & r:
                        s_unions = s_unions.union(r)
                    else:
                        outside_s.append(r)
                reduced_list_of_sets = (outside_s + [s_unions])
            return reduced_list_of_sets

        # Getting a list of subject instances (the keys of the dictionary passed to the function)
        subjects = list(subjects_markings_parameters.keys())
        # Initializing a version of the above dictionary with all values empty lists
        subjects_similar_markings = dict((subject, []) for subject in subjects)
        # Iterating through subjects
        for subject in subjects:
            # Getting the list of marking_dicts corresponding to markings that were made on this subject
            marking_dicts = subjects_markings_parameters[subject]
            # Initializing a list to track pairs of similar markings (one of whose center lies within the other's area)
            similar_markings_pairs = []
            # Iterating through all possible pairs of markings in marking_dicts;
            # checking whether the center of `inside' is within the area occupied by `outside'
            for outside in marking_dicts:
                for inside in marking_dicts:
                    """
                    NOTE: If it were true that (a and b satisfy the below) <=> (b and a satisfy the below),
                    we would only need to examine the pairs in the upper-right-triangle portion of the square grid 
                    of pairs of elements of `marking_dicts' (where `marking_dicts' is on both the horizontal
                    and vertical axes of the grid). This is not always true however (we can imagine a marking `a'
                    whose center lies on the outskirts of another marking `b', such that the center of `b' does not
                    lie in `a'), so we must examine the entire grid of pairs.
                    """
                    """
                    Performing the above described check by verifying that the value of the general ellipse equation,
                       (((x - cx) * np.cos(angle) + (y - cy) * np.sin(angle)) / xdim) ** 2 + \
                       (((x - cx) * np.sin(angle) - (y - cy) * np.cos(angle)) / ydim) ** 2
                     where (x, y) is the center of `inside' and all other parameters correspond to `outside',
                     is less than or equal to one.
                     """
                    if ellipse_eq_lhs(inside['cx'], inside['cy'], outside['cx'], outside['cy'],
                                      outside['xdim'], outside['ydim'], outside['angle']) <= 1:
                        # Appending the set {outside, inside} (in hashable form) to the list of similar pairs
                        similar_markings_pairs.append({hashDict(outside), hashDict(inside)})
            # By combining all sets that share elements, reducing the list of pairs to a list of disjoint sets
            similar_markings = reduce_list_of_sets(similar_markings_pairs)
            # Converting sets to tuples and reverted the hashed dictionaries to their original form;
            # associating with each subject a list of tuples of similar markings that were made on it
            subjects_similar_markings[subject] = [tuple([unhashDict(d) for d in s]) for s in similar_markings]
        return subjects_similar_markings

    def get_subjects_features_to_promote(self, subjects_similar_markings):
        """
        Returns subjects_features_past_positive_threshold, a dict like:
            'subject_id': [dictionaries corresponding to features whose individually-calculated probability of
                          being `positive' surpasses the promotion threshold, formatted like
                            {'positive_probability': , 'subject_history': , 'similar_marking_dicts': }
                          ]
            subjects_similar_markings: a dictionary like
                `subject_id': [tuples like (similar markings' marking_dicts)]
        """
        # Getting a list of subject instances (the keys of the dictionary passed to the function)
        subjects = list(subjects_similar_markings.keys())
        # Initializing a version of the above dictionary with all values empty lists
        subjects_features_past_positive_threshold = dict((subject, []) for subject in subjects)
        for subject in subjects:
            similar_marking_dicts_tuples = subjects_similar_markings[subject]
            for similar_marking_dicts_tuple in similar_marking_dicts_tuples:
                classification_ids = [d['classification_id'] for d in similar_marking_dicts_tuple]
                classification_ids_subject_history = \
                    self.get_classification_ids_subject_history(subject, classification_ids)

                classification_ids_user_scores = [h[2] for h in classification_ids_subject_history]
                feature_positive_probability = \
                    self.calculate_positive_probability(classification_ids_user_scores, self.positive_prior)
                if feature_positive_probability > self.promotion_threshold:
                    subjects_features_past_positive_threshold[subject].append(
                        {'positive_probability': feature_positive_probability,
                         'subject_history': classification_ids_subject_history,
                         'similar_marking_dicts': similar_marking_dicts_tuple})
        for key in subjects_features_past_positive_threshold.keys():
            if not subjects_features_past_positive_threshold[key]:
                subjects_features_past_positive_threshold.pop(key)
        return subjects_features_past_positive_threshold

    @staticmethod
    def get_classification_ids_subject_history(subject, classification_ids):
        classification_ids = [int(ID) for ID in classification_ids]
        return [sh for sh in subject.history[1:] if int(sh[0]) in classification_ids]

    def calculate_positive_probability(self, user_scores, prior_positive_probability, do_not_depreciate=True):
        positive_probability = [prior_positive_probability]
        i = 0
        for user_score in user_scores:
            if (user_tpr := user_score["1"][1]) < (user_fpr := user_score["0"][1]) and do_not_depreciate:
                # Ignore classifications from users that would decrease the `positive' probability of the feature
                continue
            positive_probability.append(self.positive_probability_update(positive_probability[i], user_score))
            i += 1
        return positive_probability[-1]

    @staticmethod
    def positive_probability_update(prior_positive_probability, user_score):
        """
        Assumes that a marking was made, and updates the subject's probability according
        to the 'score' (confusion matrix) of the user who made the classification.
            prior_positive_probability: prior probability that the subject is `positive' (contains a meltpatch)
            user_score: confusion matrix assigned to a kSWAP user;
                        {"0": [True Negative, False Positive],
                        "1": [False Negative, True Positive]}
        """
        user_tpr = user_score["1"][1]
        user_fpr = user_score["0"][1]
        prior = prior_positive_probability
        return (user_tpr * prior) / (user_tpr * prior + user_fpr * (1 - prior))

    def get_feature_to_promote_metadata_dicts(self, subjects_features_to_promote):
        """
        Create dictionaries to be written into `marking_subjects.csv' and `Marking_Manifest.xlsx' using
        the information contained within the dictionaries within the lists associated with each subject in
        `subjects_features_to_promote'.
        Returns `subjects_features_to_promote_metadata', a dictionary with key-value pairs
            `subject_id': [dictionaries with `marking_fieldnames' as keys]
        ...
            subjects_features_to_promote: a dictionary with key-value pairs
                'subject_id': [dictionaries corresponding to granite features whose individually-calculated
                              probability of being `positive' surpasses the promotion threshold, formatted like
                                {'positive_probability': , 'subject_history': , 'similar_marking_dicts': } ]
        """
        # Getting a list of subject instances (the keys of the dictionary passed to the function)
        subjects = list(subjects_features_to_promote.keys())
        # Initializing a list to contain to-be-promoted features' metadata dictionaries
        feature_to_promote_metadata_dicts = []
        # Initializing indexing variable used to assign marking subjects' unique IDs
        mrk_i = self.marking_manifest.get_first_empty_row() - 1
        for mrk_i, subject in enumerate(subjects, start=mrk_i):
            features_to_promote = subjects_features_to_promote[subject]
            for feature_to_promote in features_to_promote:
                similar_marking_dicts = feature_to_promote['similar_marking_dicts']
                similar_marking_classification_ids = [m['classification_id'] for m in similar_marking_dicts]
                similar_marking_cxs = [m['cx'] for m in similar_marking_dicts]
                similar_marking_cys = [m['cy'] for m in similar_marking_dicts]
                similar_marking_xdims = [m['xdim'] for m in similar_marking_dicts]
                similar_marking_ydims = [m['ydim'] for m in similar_marking_dicts]
                similar_marking_angles = [m['angle'] for m in similar_marking_dicts]
                similar_markings_average_cx = sum(similar_marking_cxs) / len(similar_marking_cxs)
                similar_markings_average_cy = sum(similar_marking_cys) / len(similar_marking_cys)
                similar_markings_average_xdim = sum(similar_marking_xdims) / len(similar_marking_xdims)
                similar_markings_average_ydim = sum(similar_marking_ydims) / len(similar_marking_ydims)
                similar_markings_average_angle = sum(similar_marking_angles) / len(similar_marking_angles)
                subject_clean_row = self.cleaned_classifications_csv.find_row(str(subject.subject_id),
                                                                              column_header='subject_znv_id')
                feature_to_promote_metadata = self.marking_metadata_row_template.copy()
                feature_to_promote_metadata = fill_dict_from_dict(feature_to_promote_metadata, subject_clean_row)
                marking_subject_id = 'm' + str(mrk_i)
                feature_to_promote_metadata['!subject_id'] = marking_subject_id
                feature_to_promote_metadata['#file_name'] = marking_subject_id + '_' + subject_clean_row['#file_name']
                feature_to_promote_metadata['#subject_znv_id'] = subject.subject_id
                feature_to_promote_metadata['#exp_file_name'] = subject_clean_row['#file_name']
                feature_to_promote_metadata['#resize_factor'] = eval(subject_clean_row['#resize_factor'])
                feature_to_promote_metadata['#average_cx'] = similar_markings_average_cx
                feature_to_promote_metadata['#average_cy'] = similar_markings_average_cy
                feature_to_promote_metadata['#average_xdim'] = similar_markings_average_xdim
                feature_to_promote_metadata['#average_ydim'] = similar_markings_average_ydim
                feature_to_promote_metadata['#average_angle'] = similar_markings_average_angle
                feature_to_promote_metadata['#number_of_classifications'] = len(similar_marking_dicts)
                feature_to_promote_metadata['#positive_probability'] = feature_to_promote['positive_probability']
                feature_to_promote_metadata['#classification_ids'] = similar_marking_classification_ids
                feature_to_promote_metadata['#subject_history'] = feature_to_promote['subject_history']
                feature_to_promote_metadata['#appearance_parameters'] = get_numerical_class_vars(PromoteSubjects)
                feature_to_promote_metadata_dicts.append(feature_to_promote_metadata)
        return feature_to_promote_metadata_dicts

    @staticmethod
    def retire_promoted(feature_to_promote_metadata_dicts):
        """
        Retires from Zooniverse experiment subjects in which there are promoted features.
        """
        zooniverse_subject_ids = [d['#subject_znv_id'] for d in feature_to_promote_metadata_dicts]
        zu = ZooniverseUtils(workflow_id='first')
        zu.retire_subjects(zooniverse_subject_ids)

    @staticmethod
    def get_needed_image_names(feature_to_promote_metadata_dicts):
        """
        Gets the names of all images in which the to-be-promoted features are contained.
            feature_to_promote_metadata_dicts: list of to-be-promoted features' metadata
        """
        original_image_names = [d['#pre_file_name'] for d in feature_to_promote_metadata_dicts]
        needed_image_names = []
        [needed_image_names.append(original_image_name) for original_image_name in original_image_names
         if original_image_name not in needed_image_names]
        return needed_image_names

    @staticmethod
    def correct_pre_file_name(pre_file_name):
        if '.' in pre_file_name:
            return pre_file_name
        elif pre_file_name[-4:] == 'jpeg':
            return pre_file_name.replace('jpeg', '.jpeg')
        elif pre_file_name[-3:] == 'jpg':
            return pre_file_name.replace('jpg', '.jpg')
        elif pre_file_name[-3:] == 'png':
            return pre_file_name.replace('png', '.png')

    def create_marking_images(self, feature_to_promote_metadata_dicts):
        """
        Crops the to-be-promoted features in the original (non-resized, higher-resolution) images
        and draws the `consensus marking', whose parameters are the average parameters of the
        markings made on the feature, using the to-be-promoted features' metadata
        and class variables pertaining the the consensus markings' continuity, opacity,
        color, and thickness as well as the final marking images' number of buffer pixels
         between the consensus marking and image edge and factor by which the to-be-promoted
         features' cropped images are to be enlarged.
            feature_to_promote_metadata_dicts: list of to-be-promoted features' metadata
        """
        for feature_to_promote_metadata_dict in feature_to_promote_metadata_dicts:
            original_image_name = feature_to_promote_metadata_dict['#pre_file_name']
            original_image_path = os.path.join(fetched_images_folder, original_image_name)
            original_image = cv2.imread(original_image_path)
            original_h, original_w = original_image.shape[0:2]
            marking_image_name = feature_to_promote_metadata_dict['#file_name']
            marking_image_path = os.path.join(marking_subjects_folder, marking_image_name)
            resize_factor = feature_to_promote_metadata_dict['#resize_factor']
            average_cx = int(feature_to_promote_metadata_dict['#average_cx'] / resize_factor)
            average_cy = int(feature_to_promote_metadata_dict['#average_cy'] / resize_factor)
            average_xdim = int(feature_to_promote_metadata_dict['#average_xdim'] / resize_factor)
            average_ydim = int(feature_to_promote_metadata_dict['#average_ydim'] / resize_factor)
            average_angle = feature_to_promote_metadata_dict['#average_angle']
            min_x, max_x, min_y, max_y = \
                ellipse_extrema(average_cx, average_cy, average_xdim, average_ydim, np.deg2rad(average_angle))
            feature_cropped_image_indices = np.index_exp[
                max(0, min_y - marking_edge_buffer_pixels): min(original_h, max_y + marking_edge_buffer_pixels),
                max(0, min_x - marking_edge_buffer_pixels): min(original_w, max_x + marking_edge_buffer_pixels), :]
            feature_cropped_image = original_image[feature_cropped_image_indices].copy()
            consensus_border_image = draw_dashed_ellipse(original_image.copy(), average_cx, average_cy,
                                                         average_xdim, average_ydim, average_angle,
                                                         marking_border_color, marking_border_thickness,
                                                         marking_border_gaps_angular_extent,
                                                         marking_border_segments_angular_extent)
            cropped_consensus_border_image = consensus_border_image[feature_cropped_image_indices]
            alpha = marking_border_opacity
            transparent_cropped_consensus_border_image = \
                cv2.addWeighted(feature_cropped_image, 1 - alpha, cropped_consensus_border_image, alpha, 1)
            c = marking_image_enlarge_factor
            h, w = transparent_cropped_consensus_border_image.shape[0:2]
            marking_image = cv2.resize(transparent_cropped_consensus_border_image, (c * w, c * h))
            cv2.imwrite(marking_image_path, marking_image)

    def write_metadata(self, feature_to_promote_metadata_dicts):
        for metadata_dict in feature_to_promote_metadata_dicts:
            # Verifying the each metadata dict contains the every expected and no unexpected key-value pairs
            verify_dict('marking', metadata_dict, marking_fieldnames)
        self.marking_csv.write_rows(feature_to_promote_metadata_dicts, dict_writer=True)
        self.marking_manifest.write_rows(feature_to_promote_metadata_dicts, dict_writer=True)
        self.marking_manifest_csv.write_rows(feature_to_promote_metadata_dicts, dict_writer=True)

    def upload(self):
        """
        Uploads marking subjects to Zooniverse and Google Drive; uploads manifests to Github and Google Drive.
        """
        upload_subjects_to_zooniverse(marking_csv_path, marking_subject_set_id)
        self.gd.upload_folder(marking_subjects_folder, marking_folder_drive_id)
        push_files_to_GitHub(marking_manifest_path, marking_manifest_csv_path)
        self.gd.upload_file(marking_manifest_path, manifests_folder_drive_id, replace_existing=True)
        self.gd.upload_file(marking_manifest_csv_path, manifests_csv_folder_drive_id, replace_existing=True)

    @staticmethod
    def clear_folders():
        make_folder(marking_subjects_folder, clear_existing=True)
        make_folder(fetched_images_folder, clear_existing=True)
